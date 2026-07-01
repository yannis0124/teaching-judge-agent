from __future__ import annotations

from pathlib import Path


def write_summary_ranking(output_path: Path, rows: list[dict]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as exc:
        raise RuntimeError("缺少openpyxl依赖，请先安装requirements.txt。") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "summary_ranking"
    headers = [
        "candidate_id",
        "status",
        "case_design_score_20",
        "lesson_plan_score_20",
        "live_teaching_score_60",
        "available_total_score",
        "available_max_score",
        "full_total_score_100",
        "can_score_total_100",
        "missing_sections",
        "manual_review_count",
        "three_entries_summary",
        "ai_application_type",
        "material_consistency_summary",
        "teaching_closure_summary",
        "key_review_points",
        "remarks",
    ]
    merged_rows = _merge_summary_rows(output_path, headers, rows, load_workbook)
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in sorted(merged_rows, key=_summary_sort_key, reverse=True):
        sheet.append([row.get(header, "") for header in headers])
    for column_cells in sheet.columns:
        max_len = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 12), 50)
        for cell in column_cells:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def _merge_summary_rows(output_path: Path, headers: list[str], rows: list[dict], load_workbook) -> list[dict]:
    merged: dict[str, dict] = {}
    if output_path.exists():
        workbook = load_workbook(output_path, data_only=True, read_only=True)
        try:
            sheet = workbook["summary_ranking"] if "summary_ranking" in workbook.sheetnames else workbook.active
            existing_headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            for values in sheet.iter_rows(min_row=2, values_only=True):
                row = {
                    header: "" if value is None else value
                    for header, value in zip(existing_headers, values)
                    if header in headers
                }
                candidate_id = str(row.get("candidate_id") or "").strip()
                if candidate_id:
                    row["candidate_id"] = candidate_id
                    merged[candidate_id] = row
        finally:
            workbook.close()

    for row in rows:
        candidate_id = str(row.get("candidate_id") or "").strip()
        if not candidate_id:
            continue
        merged[candidate_id] = {header: row.get(header, "") for header in headers}
        merged[candidate_id]["candidate_id"] = candidate_id

    return [{header: row.get(header, "") for header in headers} for row in merged.values()]


def _summary_sort_key(row: dict) -> float:
    try:
        return float(row.get("available_total_score") or -1)
    except (TypeError, ValueError):
        return -1.0


def build_summary_row(candidate_id: str, evidence_package: dict, agent_results: dict | None, note: str | None = None) -> dict:
    completeness = evidence_package.get("material_completeness", {})
    if not agent_results:
        missing_sections = _missing_sections(completeness)
        return {
            "candidate_id": candidate_id,
            "status": "unscored",
            "case_design_score_20": "",
            "lesson_plan_score_20": "",
            "live_teaching_score_60": "",
            "available_total_score": "",
            "available_max_score": completeness.get("available_max_score", ""),
            "full_total_score_100": "",
            "can_score_total_100": completeness.get("can_score_total_100", False),
            "missing_sections": ",".join(missing_sections),
            "manual_review_count": len(evidence_package.get("material_integrity", {}).get("notes", [])),
            "three_entries_summary": "",
            "ai_application_type": "",
            "material_consistency_summary": "",
            "teaching_closure_summary": "",
            "key_review_points": "",
            "remarks": note or "未生成正式评分建议。",
        }
    final = agent_results.get("final_judgement", {})
    overall = agent_results.get("overall_review", {})
    dimensions = overall.get("dimensions", {}) if isinstance(overall, dict) else {}
    manual_points = []
    for item in agent_results.get("scoring_results", {}).values():
        manual_points.extend(item.get("manual_review_points", []))
    for key in ["consistency_review", "bias_review", "final_judgement", "overall_review"]:
        manual_points.extend(agent_results.get(key, {}).get("manual_review_points", []))
    manual_points = _unique_points(manual_points)
    status = "unscored" if final.get("available_total_score") in (None, "") else ("needs_review" if manual_points else "scored")
    review_focus = dimensions.get("evidence_limitations_and_review_focus", {})
    return {
        "candidate_id": candidate_id,
        "status": status,
        "case_design_score_20": final.get("case_design_score_20"),
        "lesson_plan_score_20": final.get("lesson_plan_score_20"),
        "live_teaching_score_60": final.get("live_teaching_score_60"),
        "available_total_score": final.get("available_total_score"),
        "available_max_score": final.get("available_max_score"),
        "full_total_score_100": final.get("full_total_score_100") if final.get("can_score_total_100") else "",
        "can_score_total_100": final.get("can_score_total_100"),
        "missing_sections": ",".join(final.get("missing_sections", [])),
        "manual_review_count": len(manual_points),
        "three_entries_summary": dimensions.get("three_entries_integration", {}).get("summary", ""),
        "ai_application_type": dimensions.get("ai_application_effectiveness", {}).get("application_type", ""),
        "material_consistency_summary": dimensions.get("material_consistency", {}).get("summary", ""),
        "teaching_closure_summary": dimensions.get("teaching_closure", {}).get("summary", ""),
        "key_review_points": _join(
            review_focus.get("high_priority_review_points", [])
            or final.get("manual_review_points", [])
            or []
        ),
        "remarks": note or "",
    }


def _missing_sections(completeness: dict) -> list[str]:
    missing = []
    if not completeness.get("can_score_case_design_20"):
        missing.append("case_design")
    if not completeness.get("can_score_lesson_plan_20"):
        missing.append("lesson_plan")
    if not completeness.get("can_score_live_teaching_60"):
        missing.append("live_teaching")
    return missing


def _join(values: list | str | None) -> str:
    if values is None:
        return ""
    if isinstance(values, str):
        return values
    return "\n".join(str(value) for value in values)


def _unique_points(values: list) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        text = str(value).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result
